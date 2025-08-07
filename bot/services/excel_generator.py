import openpyxl
import shutil
import os

def generate_report(report_data, output_path="data/generated_report.xlsx"):
    template_path = os.path.join('data', 'report_template.xlsx')
    report_path = output_path

    shutil.copy(template_path, report_path)
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active

    # 1. Организация (A6)
    ws['A6'] = report_data.get("employee_organisation", "")

    # 2. ФИО (K22)
    fio = f"{report_data.get('surname', '')} {report_data.get('name', '')} {report_data.get('patronymic_name', '')}".strip()
    ws["K22"] = fio

    # 3. Табельный номер (AR22)
    ws["AR22"] = report_data.get("employee_id", "")

    # 4. Приказ (S67)
    order = report_data.get("order", {})
    row_number = 1  # Начинаем нумерацию с 1
    if order:  # Проверяем, что order не None
        ws["A67"] = row_number  # Нумерация в столбце A
        ws["S67"] = "Приказ"  # <-- только слово "Приказ"
        ws["D67"] = order.get('order_date', '')  # Дата приказа
        ws["J67"] = order.get('order_number', '')  # Вставка номера приказа в J67
        duration = int(order.get("duration", 0) or 0)
        row_number += 1
    else:
        ws["S67"] = "Приказ не указан"
        ws["D67"] = ""
        ws["J67"] = ""  # Если приказ не указан, ячейка пустая
        duration = 0

    # 6. Билеты (D68, J68, AA68 и далее)
    tickets = report_data.get("tickets", [])
    start_row = 68
    for i, ticket in enumerate(tickets):
        row = start_row + i
        ws[f"A{row}"] = row_number  # Нумерация в столбце A
        ws[f"D{row}"] = ticket.get("ticket_date", "")
        ws[f"J{row}"] = ticket.get("ticket_number", "")
        ws[f"AA{row}"] = float(ticket.get("ticket_price", 0) or 0)
        row_number += 1
        
    # 6.1. Дата первого билета в D68
    if tickets and tickets[0].get("ticket_date"):
        ws["D68"] = tickets[0].get("ticket_date", "") if tickets else ""
    else:
        ws["D68"] = ""

    # 6.2. Слово "Билет" в S68
    if tickets:
        ws["S68"] = "Билет"

    # 5. Срок командировки (S69, AA69)
    if duration > 0:  # Только если есть срок командировки
        ws[f"A{68 + len(tickets)}"] = row_number  # Нумерация для строки суточных
        ws["S69"] = f"700 * {duration}"
        ws["AA69"] = duration * 700
        row_number += 1

    # 7. Слово "Суточные" (J69)
    if duration > 0:  # Только если есть срок командировки
        ws["J69"] = "Суточные"

    # 8. Итоговая сумма в AA79 (сумма AA67:AA78)
    ws["AA79"] = "=SUM(AA67:AA78)"

    # 9. ФИО пользователя в AK81
    fio_user = f"{report_data.get('surname', '')} {report_data.get('name', '')} {report_data.get('patronymic_name', '')}".strip()
    ws["AK81"] = fio_user

    wb.save(report_path)

if __name__ == "__main__":
    # Пример тестовых данных
    report_data = {
        "employee_organisation": "ООО АЛРОСА информационные технологии",
        "surname": "Иванов",
        "name": "Иван",
        "patronymic_name": "Иванович",
        "employee_id": "D8200684",
        "order": {
            "order_number": "123/К",
            "order_date": "01.06.2024",
            "duration": 5
        },
        "tickets": [
            {
                "ticket_date": "02.06.2024",
                "ticket_number": "1234567890",
                "ticket_price": 15000.0
            },
            {
                "ticket_date": "03.06.2024",
                "ticket_number": "0987654321",
                "ticket_price": 12000.0
            }
        ]
    }
    generate_report(report_data)
    order = report_data.get("order", {})
    tickets = report_data.get("tickets", [])
    print("Отчет сгенерирован и сохранен в data/generated_report.xlsx")
    print("order_date для отчета:", order.get('order_date', ''))
    print("ticket_date для отчета:", tickets[0].get("ticket_date", "") if tickets else None)